import os
from flask import Flask, request, jsonify, render_template, redirect
from flask_cors import CORS
import mysql.connector
import bcrypt
import random
import smtplib
from email.message import EmailMessage
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
import uuid
from functools import wraps
from flask import session as flask_session, Response
import io
import csv
from decimal import Decimal, InvalidOperation
from calendar import monthrange
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import jwt
import jinja2

load_dotenv()

app = Flask(__name__, static_folder='public', static_url_path='/public')
app.secret_key = os.environ.get('SECRET_KEY', 'default-dev-key-change-in-prod')
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
CORS(app, supports_credentials=True)

# Multi-directory template loader: supports standard templates/ and root admin_templates/
template_dirs = [
    os.path.join(app.root_path, 'templates'),
    os.path.join(app.root_path, 'admin_templates')
]
app.jinja_loader = jinja2.FileSystemLoader(template_dirs)

DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'localhost'),
    'port': int(os.environ.get('DB_PORT', 3306)),
    'user': os.environ.get('DB_USER', 'root'),
    'password': os.environ.get('DB_PASS', ''),
    'database': os.environ.get('DB_NAME', 'spxbank')
}

db_pool = mysql.connector.pooling.MySQLConnectionPool(
    pool_name="spxbank_pool",
    pool_size=5,
    pool_reset_session=True,
    **DB_CONFIG
)

def get_db_connection():
    return db_pool.get_connection()

def run_migrations():
    """Auto-migrate DB schema on startup. Safe to run repeatedly."""
    migrations = [
        # Add action column to otps if it doesn't already exist
        "ALTER TABLE otps ADD COLUMN action VARCHAR(50) NOT NULL DEFAULT 'LOGIN'",
        # transactions table for fund transfer / history feature
        """CREATE TABLE IF NOT EXISTS transactions (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            type VARCHAR(30) NOT NULL,
            amount DECIMAL(15,2) NOT NULL,
            counterparty_account VARCHAR(50),
            counterparty_name VARCHAR(255),
            note VARCHAR(255),
            balance_after DECIMAL(15,2) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )""",
        # Widen type from the original ENUM so admin fund adjustments
        # (ADMIN_CREDIT / ADMIN_DEBIT) can be stored in the same column
        "ALTER TABLE transactions MODIFY COLUMN type VARCHAR(30) NOT NULL",
        # Extra columns the admin panel's transaction views expect
        "ALTER TABLE transactions ADD COLUMN description TEXT",
        "ALTER TABLE transactions ADD COLUMN reference_id VARCHAR(50)",
        "ALTER TABLE transactions ADD COLUMN status VARCHAR(20) DEFAULT 'COMPLETED'",

        # Customer identity/profile enhancements
        "ALTER TABLE users ADD COLUMN mid_number VARCHAR(11) UNIQUE",
        "ALTER TABLE users ADD UNIQUE KEY uq_users_account_number (account_number)",
        # Additional KYC/profile information is kept separately from login/account credentials.
        """CREATE TABLE IF NOT EXISTS add_info (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL UNIQUE,
            date_of_birth DATE NULL,
            mobile_number VARCHAR(20) NULL,
            pan VARCHAR(20) NULL,
            father_name VARCHAR(255) NULL,
            alternate_email VARCHAR(255) NULL,
            communication_address TEXT NULL,
            permanent_address TEXT NULL,
            marital_status VARCHAR(30) NULL,
            religion VARCHAR(80) NULL,
            category VARCHAR(80) NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )""",
        # --- Admin system (merged from teammate's admin project) ---
        "ALTER TABLE users ADD COLUMN account_status VARCHAR(20) DEFAULT 'ACTIVE'",
        "ALTER TABLE users ADD COLUMN last_login TIMESTAMP NULL",
        """CREATE TABLE IF NOT EXISTS admin_audit_logs (
            id INT AUTO_INCREMENT PRIMARY KEY,
            admin_email VARCHAR(255) NOT NULL,
            action VARCHAR(100) NOT NULL,
            target_type VARCHAR(50),
            target_id VARCHAR(100),
            details TEXT,
            prev_value TEXT,
            new_value TEXT,
            status VARCHAR(20) DEFAULT 'SUCCESS',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS loans (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            loan_type VARCHAR(50) NOT NULL,
            amount DECIMAL(15, 2) NOT NULL,
            interest_rate DECIMAL(5, 2) NOT NULL,
            tenure_months INT NOT NULL,
            emi DECIMAL(10, 2),
            status VARCHAR(20) DEFAULT 'PENDING',
            purpose VARCHAR(255),
            employment_type VARCHAR(50),
            monthly_income DECIMAL(15, 2),
            existing_emi DECIMAL(15, 2) DEFAULT 0,
            admin_notes TEXT,
            applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            approved_at TIMESTAMP NULL,
            disbursed_at TIMESTAMP NULL,
            outstanding_principal DECIMAL(15, 2) DEFAULT 0,
            next_emi_date DATE NULL,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )""",
        "ALTER TABLE loans ADD COLUMN purpose VARCHAR(255)",
        "ALTER TABLE loans ADD COLUMN employment_type VARCHAR(50)",
        "ALTER TABLE loans ADD COLUMN monthly_income DECIMAL(15,2)",
        "ALTER TABLE loans ADD COLUMN existing_emi DECIMAL(15,2) DEFAULT 0",
        "ALTER TABLE loans ADD COLUMN disbursed_at TIMESTAMP NULL",
        "ALTER TABLE loans ADD COLUMN outstanding_principal DECIMAL(15,2) DEFAULT 0",
        "ALTER TABLE loans ADD COLUMN next_emi_date DATE NULL",
        "ALTER TABLE loans ADD COLUMN reference_id VARCHAR(50) NULL UNIQUE",
        """CREATE TABLE IF NOT EXISTS loan_payments (
            id INT AUTO_INCREMENT PRIMARY KEY,
            loan_id INT NOT NULL,
            user_id INT NOT NULL,
            amount DECIMAL(15,2) NOT NULL,
            principal_component DECIMAL(15,2) NOT NULL,
            interest_component DECIMAL(15,2) NOT NULL,
            balance_after DECIMAL(15,2) NOT NULL,
            reference_id VARCHAR(50) UNIQUE NOT NULL,
            paid_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (loan_id) REFERENCES loans(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )""",
        """CREATE TABLE IF NOT EXISTS cards (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            card_type VARCHAR(30) NOT NULL,
            card_number_last4 VARCHAR(4) NOT NULL,
            card_number_masked VARCHAR(25),
            status VARCHAR(20) DEFAULT 'REQUESTED',
            issued_at TIMESTAMP NULL,
            expires_at DATE NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )""",
        # fund_transfer defaults to FALSE — this is the actual "admin must
        # verify user before they can transfer money" gate. Everything else
        # defaults TRUE since only fund transfer was asked to be restricted.
        """CREATE TABLE IF NOT EXISTS user_privileges (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL UNIQUE,
            online_banking BOOLEAN DEFAULT TRUE,
            fund_transfer BOOLEAN DEFAULT TRUE,
            card_access BOOLEAN DEFAULT TRUE,
            loan_application BOOLEAN DEFAULT TRUE,
            high_value_transfer BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        )""",
    ]
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for sql in migrations:
            try:
                cursor.execute(sql)
                conn.commit()
            except Exception as e:
                # 1060 = Duplicate column name — column already exists, safe to skip
                if hasattr(e, 'errno') and e.errno == 1060:
                    pass
                else:
                    pass
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"[MIGRATION ERROR] Could not connect for migrations: {e}")

run_migrations()

# Backfill MID numbers and profile rows for users created before the enhancement.
def backfill_customer_identity():
    try:
        conn=get_db_connection(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM users WHERE mid_number IS NULL OR mid_number='' ")
        rows=cursor.fetchall()
        for row in rows:
            mid=None
            for _ in range(100):
                candidate=str(random.randint(10_000_000_000,99_999_999_999))
                cursor.execute("SELECT id FROM users WHERE mid_number=%s", (candidate,))
                if not cursor.fetchone():
                    mid=candidate; break
            if not mid: raise RuntimeError('Unable to backfill MID')
            cursor.execute("UPDATE users SET mid_number=%s WHERE id=%s", (mid,row['id']))
        cursor.execute("SELECT id FROM users")
        for row in cursor.fetchall():
            cursor.execute("INSERT IGNORE INTO add_info (user_id) VALUES (%s)", (row['id'],))
            cursor.execute("INSERT IGNORE INTO user_privileges (user_id, online_banking, fund_transfer, card_access, loan_application, high_value_transfer) VALUES (%s, TRUE, TRUE, TRUE, TRUE, FALSE)", (row['id'],))
        conn.commit(); cursor.close(); conn.close()
    except Exception as e:
        print(f"[IDENTITY BACKFILL WARN] {e}")

backfill_customer_identity()

def get_email_template(action, first_name, otp):
    subject = "Login Verification OTP"
    heading = "Verify your SPX Bank login"
    body_desc = "Use the OTP below to complete your sign-in. This code is valid for 1 minute."
    security_note = "🔒 Never share this OTP with anyone, including SPX Bank staff."
    footer_text = "If you didn't request this code, you can safely ignore this email."

    if action == 'REGISTER':
        subject = "Registration Verification OTP"
        heading = "Verify your email address"
        body_desc = "Use the OTP below to verify your email and complete your SPX Bank account registration. This code is valid for 2 minutes."
        security_note = "🔒 Never share this OTP with anyone, including SPX Bank staff."
        footer_text = "If you didn't attempt to create an account with SPX Bank, please ignore this email."
    elif action == 'RESET_PASSWORD':
        subject = "Password Reset OTP"
        heading = "Reset your SPX Bank password"
        body_desc = "We received a request to reset your netbanking password. Use the code below to proceed. Valid for 2 minutes."
        security_note = "🔒 SPX Bank will never ask for this code. Do not share it with anyone."
        footer_text = "If you didn't request a password reset, please ignore this email or secure your account."
    elif action == 'LOAN_VERIFY':
        subject = "Verification OTP - SPX Bank"
        heading = "Verify your SPX Bank Account"
        body_desc = "Use the OTP below to complete your verification request."
        security_note = "🔒 Never share this OTP with anyone, including SPX Bank staff."
        footer_text = "If you didn't request this verification, please ignore this email."

    plain_text = f"{heading}\n\n{body_desc}\n\nVerification Code: {otp}\n\n{security_note}"

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
    </head>
    <body style="margin: 0; padding: 0; background-color: #f8f9fa; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
        <span style="display:none; font-size:1px; color:#ffffff; line-height:1px; max-height:0px; max-width:0px; opacity:0; overflow:hidden;">
            Your SPX Bank verification code is enclosed. Please do not share this code with anyone.
        </span>
        <div style="background-color: #f8f9fa; padding: 40px 20px;">
            <div style="max-width: 500px; margin: 0 auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.05); padding: 40px; border: 1px solid #eaeaea;">
                <h1 style="color: #111827; font-size: 20px; font-weight: 700; margin-top: 0; margin-bottom: 16px; text-align: center;">
                    {heading}
                </h1>
                
                <p style="color: #4b5563; font-size: 15px; line-height: 1.6; margin-bottom: 30px; text-align: center;">
                    {body_desc}
                </p>
                
                <div style="background-color: #f4f5f7; border: 1px solid #eaeaea; padding: 24px; text-align: center; border-radius: 6px; margin-bottom: 30px;">
                    <div style="font-size: 32px; font-weight: 700; color: #5C2D91; letter-spacing: 6px; margin-left: 8px;">
                        {otp}
                    </div>
                </div>
                
                <p style="color: #6b7280; font-size: 13px; line-height: 1.5; margin-bottom: 0; text-align: center;">
                    {security_note}
                </p>
                
                <hr style="border: none; border-top: 1px solid #eaeaea; margin: 30px 0;">
                
                <div style="text-align: center; color: #9ca3af; font-size: 12px; line-height: 1.5;">
                    <p style="margin: 0 0 8px 0;">{footer_text}</p>
                    <p style="margin: 0;">&copy; 2026 SPX Bank. All rights reserved.</p>
                    <p style="font-size: 11px; color: #888888; margin-top: 15px;">
                        <span style="display:none; font-size:1px; color:#ffffff; opacity:0;">Ref: {uuid.uuid4()}</span>
                    </p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    return subject, html_content, plain_text

def send_real_email(to_email, subject, html_body, plain_text):
    try:
        smtp_server = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.environ.get('SMTP_PORT', 465))
        smtp_user = os.environ.get('SMTP_USER')
        smtp_pass = os.environ.get('SMTP_PASS')

        if not smtp_user or not smtp_pass:
            print("WARNING: SMTP credentials not set. Simulated email:")
            print(f"To: {to_email}\nSubject: {subject}\nBody: HTML Content rendered")
            return True

        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = f"SPX Bank <{smtp_user}>"
        msg['To'] = to_email
        msg.set_content(plain_text)
        msg.add_alternative(html_body, subtype='html')

        # Logo attachment logic removed

        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

def login_required(f):
    """Guards API routes and View routes that require a real logged-in server-side session."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in flask_session:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': 'Session expired. Please log in again.'}), 401
            else:
                return redirect('/registration/welcome')
        return f(*args, **kwargs)
    return wrapper


# ==========================================================================
# ADMIN SYSTEM (merged from teammate's admin project)
# ==========================================================================
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'admin@spxbank.local')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'change-this-admin-password')
ADMIN_NAME = os.environ.get('ADMIN_NAME', 'Bank Administrator')

def generate_admin_token():
    now_time = datetime.now(timezone.utc)
    payload = {
        'sub': ADMIN_EMAIL,
        'name': ADMIN_NAME,
        'role': 'MASTER_ADMIN',
        'iat': now_time,
        'exp': now_time + timedelta(hours=12)
    }
    return jwt.encode(payload, app.secret_key, algorithm='HS256')

def require_admin_token(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
        elif request.args.get('token'):
            token = request.args.get('token')

        if not token:
            return jsonify({'success': False, 'message': 'Admin authorization token required'}), 401

        try:
            payload = jwt.decode(token, app.secret_key, algorithms=['HS256'])
            if payload.get('role') != 'MASTER_ADMIN' or payload.get('sub') != ADMIN_EMAIL:
                return jsonify({'success': False, 'message': 'Unauthorized admin privileges'}), 403
            request.admin_user = payload
        except jwt.ExpiredSignatureError:
            return jsonify({'success': False, 'message': 'Admin session expired. Please log in again.'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'success': False, 'message': 'Invalid admin authorization token.'}), 401

        return f(*args, **kwargs)
    return decorated

def log_admin_action(action, target_type=None, target_id=None, details=None, prev_value=None, new_value=None, status='SUCCESS'):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        sql = """
        INSERT INTO admin_audit_logs (admin_email, action, target_type, target_id, details, prev_value, new_value, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (ADMIN_EMAIL, action, target_type, str(target_id) if target_id else None, details, str(prev_value) if prev_value else None, str(new_value) if new_value else None, status))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"[AUDIT LOG ERROR] {e}")


# --------------------------------------------------------------------------
# CUSTOMER DATA HELPERS
# --------------------------------------------------------------------------
def generate_unique_mid(cursor):
    """Generate a unique 11-digit customer/member identification number."""
    for _ in range(100):
        mid = str(random.randint(10_000_000_000, 99_999_999_999))
        cursor.execute("SELECT id FROM users WHERE mid_number=%s", (mid,))
        if not cursor.fetchone():
            return mid
    raise RuntimeError("Unable to generate a unique MID number")


def get_customer_context(user_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT u.id, u.username, u.email, u.first_name, u.last_name,
                   u.account_number, u.balance, u.account_status, u.last_login,
                   u.mid_number, u.created_at,
                   ai.date_of_birth, ai.mobile_number, ai.pan, ai.father_name,
                   ai.alternate_email, ai.communication_address,
                   ai.permanent_address, ai.marital_status, ai.religion, ai.category
            FROM users u
            LEFT JOIN add_info ai ON ai.user_id=u.id
            WHERE u.id=%s
        """, (user_id,))
        user = cursor.fetchone()
        if not user:
            return None
        cursor.execute("SELECT fund_transfer, online_banking FROM user_privileges WHERE user_id=%s", (user_id,))
        priv = cursor.fetchone() or {}
        user['fund_transfer_enabled'] = bool(priv.get('fund_transfer', False))
        user['online_banking_enabled'] = bool(priv.get('online_banking', True))
        user['balance'] = f"{Decimal(str(user['balance'] or 0)):,.2f}"
        user['name'] = f"{user['first_name']} {user['last_name']}".strip()
        user['initials'] = ''.join(part[0] for part in user['name'].split() if part)[:2].upper()
        if user.get('last_login'):
            user['last_login_display'] = user['last_login'].strftime('%d %b %Y, %I:%M %p')
        else:
            user['last_login_display'] = 'First login'
        if user.get('date_of_birth'):
            user['date_of_birth'] = user['date_of_birth'].isoformat()
        if user.get('created_at'):
            user['created_at'] = user['created_at'].strftime('%d %b %Y, %I:%M %p')
        return user
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()


def get_statement_range(duration=None, financial_year=None, start_date=None, end_date=None):
    today = datetime.now().date()
    if start_date and end_date:
        return start_date, end_date
    if financial_year:
        try:
            start_year = int(financial_year[:4])
            return datetime(start_year, 4, 1).date(), datetime(start_year + 1, 3, 31).date()
        except (ValueError, TypeError):
            pass
    duration = (duration or 'CURRENT_MONTH').upper().replace(' ', '_')
    if duration == 'CURRENT_MONTH':
        return today.replace(day=1), today
    if duration == 'LAST_MONTH':
        first_current = today.replace(day=1)
        last_prev = first_current - timedelta(days=1)
        return last_prev.replace(day=1), last_prev
    if duration == 'LAST_3_MONTHS':
        first_current = today.replace(day=1)
        cursor = first_current
        for _ in range(2):
            cursor = (cursor.replace(day=1) - timedelta(days=1)).replace(day=1)
        return cursor, today
    if duration == 'CURRENT_FINANCIAL_YEAR':
        fy_start_year = today.year if today.month >= 4 else today.year - 1
        return datetime(fy_start_year, 4, 1).date(), today
    if duration == 'LAST_FINANCIAL_YEAR':
        fy_start_year = today.year if today.month >= 4 else today.year - 1
        return datetime(fy_start_year - 1, 4, 1).date(), datetime(fy_start_year, 3, 31).date()
    # Custom range is handled by explicit start/end dates.
    return today.replace(day=1), today


def fetch_transactions_for_range(user_id, start_date, end_date):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, type, amount, counterparty_account, counterparty_name,
                   note, description, reference_id, status, balance_after, created_at
            FROM transactions
            WHERE user_id=%s AND DATE(created_at) BETWEEN %s AND %s
            ORDER BY created_at DESC, id DESC
        """, (user_id, start_date, end_date))
        return cursor.fetchall()
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()


def transaction_payload(rows):
    labels = {
        'TRANSFER_OUT': 'Money Sent', 'TRANSFER_IN': 'Money Received',
        'DEPOSIT': 'Deposit', 'WITHDRAWAL': 'Withdrawal',
        'ADMIN_CREDIT': 'Admin Credit', 'ADMIN_DEBIT': 'Admin Debit',
        'CREDIT': 'Credit', 'DEBIT': 'Debit'
    }
    result=[]
    for r in rows:
        incoming = r['type'] in ('TRANSFER_IN','DEPOSIT','ADMIN_CREDIT', 'CREDIT')
        
        t_type = r['type']
        desc = r.get('description') or ''
        note = r.get('note') or ''
        c_name = r.get('counterparty_name') or 'Unknown'
        c_acc = r.get('counterparty_account') or ''
        c_acc_last4 = c_acc[-4:] if len(c_acc) >= 4 else c_acc
        ref_id = r.get('reference_id') or f"TXN-{r['id']}"
        
        if t_type == 'CREDIT' and ('Loan' in note or 'LOAN' in desc):
            title = 'Loan Disbursal - Personal Loan'
            subtext = f'Ref: {ref_id} • SPX Bank'
        elif t_type == 'TRANSFER_OUT':
            title = f'Transfer to {c_name}'
            subtext = f'SPX Transfer • A/C ••••{c_acc_last4}' if c_acc_last4 else 'SPX Transfer'
        elif t_type == 'TRANSFER_IN':
            title = f'Transfer from {c_name}'
            subtext = f'SPX Transfer • A/C ••••{c_acc_last4}' if c_acc_last4 else 'SPX Transfer'
        elif t_type == 'DEBIT' and ('EMI' in note.upper() or 'LOAN' in desc.upper()):
            title = 'EMI Repayment - Personal Loan'
            subtext = f'Auto-Debit • Ref: {ref_id}'
        elif t_type == 'ADMIN_CREDIT':
            title = 'Account Credit - Adjustment'
            subtext = f'SPX Desk • Ref: SPXAD{r["id"]}'
        else:
            title = desc if desc else 'Net-Banking Transaction'
            subtext = f'SPX Net-Banking • Ref: SPX-{r["id"]}'

        result.append({
            'id': r['id'], 'type': r['type'], 'typeLabel': labels.get(r['type'], r['type']),
            'title': title, 'subtext': subtext,
            'amount': f"{Decimal(str(r['amount'] or 0)):,.2f}",
            'direction': 'IN' if incoming else 'OUT',
            'counterpartyAccount': r.get('counterparty_account'),
            'counterpartyName': r.get('counterparty_name'),
            'note': r.get('note'), 'description': r.get('description'),
            'referenceId': r.get('reference_id') or f"TXN-{r['id']}",
            'status': r.get('status') or 'COMPLETED',
            'balanceAfter': f"{Decimal(str(r['balance_after'] or 0)):,.2f}",
            'date': r['created_at'].strftime('%d %b %Y, %I:%M %p') if r.get('created_at') else ''
        })
    return result

# --- VIEW ROUTES ---
@app.route('/')
def root():
    """Redirect root to canonical welcome URL."""
    return redirect('/registration/welcome')

@app.route('/registration/welcome')
def index():
    """Entry point: Login & Registration SPA."""
    return render_template('index.html')

@app.route('/registration/account-identification')
def password_recovery():
    """Password Recovery landing — serves the same SPA with the forgot-password view."""
    return render_template('index.html')

@app.route('/home/landingPage/homePage')
@login_required
def overview():
    user = get_customer_context(flask_session['user_id'])
    if not user or not user['online_banking_enabled'] or user['account_status'] != 'ACTIVE':
        flask_session.clear()
        return redirect('/registration/welcome')
    return render_template('overview.html', user=user, user_initials=user['initials'], account_number=user['account_number'], balance=user['balance'])

@app.route('/logout')
def logout_redirect():
    """Server-side logout: clears the real session and sends the client back to login."""
    flask_session.clear()
    return redirect('/registration/welcome')

@app.route('/api/logout', methods=['POST'])
def api_logout():
    flask_session.clear()
    return jsonify({'success': True})

@app.route('/home/landingPage/manageRelationship/transactionAccounts')
@login_required
def accounts():
    user = get_customer_context(flask_session['user_id'])
    if not user or user['account_status'] != 'ACTIVE':
        return redirect('/registration/welcome')
    return render_template('accounts.html', user=user, user_initials=user['initials'], account_number=user['account_number'], balance=user['balance'])

@app.route('/home/landingPage/statement')
@login_required
def statement_redirect():
    return redirect('/home/landingPage/manageRelationship/transactionAccounts')

@app.route('/home/landingPage/profilePage/services/manageProfile/personalDetails')
@login_required
def profile_page():
    user = get_customer_context(flask_session['user_id'])
    if not user or user['account_status'] != 'ACTIVE':
        return redirect('/registration/welcome')
    return render_template('profile.html', user=user, user_initials=user['initials'])

@app.route('/home/landingPage/profilePage/send-money/fund-transfer')
@login_required
def send_money_page():
    user = get_customer_context(flask_session['user_id'])
    if not user or user['account_status'] != 'ACTIVE':
        return redirect('/registration/welcome')
    return render_template('payments/send_money.html', user=user, user_initials=user['initials'], account_number=user['account_number'], balance=user['balance'])
# -------------------

@app.route('/home/landingPage/loans/manage', methods=['GET'])
@login_required
def loans_manage():
    user = get_customer_context(flask_session['user_id'])
    if not user or user['account_status'] != 'ACTIVE':
        return redirect('/registration/welcome')
    return render_template('loans/loans_manage.html', user=user, user_initials=user['initials'], account_number=user['account_number'], balance=user['balance'])
# -------------------

@app.route('/home/landingPage/loans/personal-loan/')
@login_required
def personal_loan_landing_page():
    user = get_customer_context(flask_session['user_id'])
    if not user or user['account_status'] != 'ACTIVE':
        return redirect('/registration/welcome')
    return render_template('loans/loans_personal.html', user=user, user_initials=user['initials'], account_number=user['account_number'], balance=user['balance'])

@app.route('/home/landingPage/loans/personal-loan/apply/')
@app.route('/home/landingPage/loans/personal-loan/basicinfo/email-verify/')
@app.route('/home/landingPage/loans/personal-loan/doc-upload/')
@app.route('/home/landingPage/loans/personal-loan/apply-loan-application/')
@app.route('/home/landingPage/loans/personal-loan/preview-loan-application/')
@app.route('/home/landingPage/loans/personal-loan/application-submit-status/')
@login_required
def personal_loan_apply_page():
    user = get_customer_context(flask_session['user_id'])
    if not user or user['account_status'] != 'ACTIVE':
        return redirect('/registration/welcome')
    return render_template('loans/loans_personal_apply.html', user=user, user_initials=user['initials'], account_number=user['account_number'], balance=user['balance'])

@app.route('/api/loans/send-otp', methods=['POST'])
@login_required
def send_loan_otp():
    user = get_customer_context(flask_session['user_id'])
    if not user:
        return jsonify({'status': 'error', 'message': 'User not found'}), 404
        
    email = user['email']
    otp = str(random.randint(100000, 999999))
    flask_session['loan_flow_otp'] = otp
    
    subject, html_body, plain_text = get_email_template('LOAN_VERIFY', user['first_name'], otp)
    send_real_email(email, subject, html_body, plain_text)
    
    return jsonify({'status': 'success', 'message': 'OTP sent'})

@app.route('/api/loans/verify-otp', methods=['POST'])
@login_required
def verify_loan_otp():
    data = request.json
    submitted_otp = data.get('otp')
    saved_otp = flask_session.get('loan_flow_otp')
    
    if saved_otp and str(submitted_otp) == str(saved_otp):
        flask_session['loan_flow_verified'] = True
        return jsonify({'status': 'success'})
    else:
        return jsonify({'status': 'error', 'message': 'Invalid OTP'}), 400
# -------------------

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')
    first_name = data.get('firstName')
    last_name = data.get('lastName')

    if not all([username, email, password, first_name, last_name]):
        return jsonify({'success': False, 'message': 'Missing fields'}), 400

    hashed_pw = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # 1. Enforce OTP validation
        cursor.execute("SELECT * FROM otps WHERE email=%s AND action='REGISTER' AND used=TRUE ORDER BY created_at DESC LIMIT 1", (email,))
        otp_record = cursor.fetchone()
        
        if not otp_record:
            return jsonify({'success': False, 'message': 'OTP verification required'}), 403
            
        if otp_record['expires_at'] < datetime.now():
            return jsonify({'success': False, 'message': 'Verified OTP has expired. Please request a new one.'}), 403

        # 2. Duplicate check
        cursor.execute("SELECT id FROM users WHERE username=%s OR email=%s", (username, email))
        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'Username or email already exists'}), 409
            
        # 3. Consume OTP
        cursor.execute("DELETE FROM otps WHERE id=%s", (otp_record['id'],))
        
        account_number = None
        for _ in range(100):
            candidate = f"884901{random.randint(100000, 999999)}"
            cursor.execute("SELECT id FROM users WHERE account_number=%s", (candidate,))
            if not cursor.fetchone():
                account_number = candidate
                break
        if not account_number:
            raise RuntimeError('Unable to generate unique account number')
        mid_number = generate_unique_mid(cursor)

        sql = """
        INSERT INTO users (username, email, password_hash, first_name, last_name, account_number, mid_number, balance) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (username, email, hashed_pw, first_name, last_name, account_number, mid_number, 25000.00))
        new_user_id = cursor.lastrowid

        # New users start WITH fund transfer access enabled by default.
        # High value transfer remains disabled by default.
        cursor.execute(
            """INSERT INTO user_privileges (user_id, online_banking, fund_transfer, card_access, loan_application, high_value_transfer)
               VALUES (%s, TRUE, TRUE, TRUE, TRUE, FALSE)""",
            (new_user_id,)
        )
        cursor.execute("INSERT INTO add_info (user_id) VALUES (%s)", (new_user_id,))
        conn.commit()

        # Let the admin know a new account needs verification
        try:
            admin_subject = "New User Registration — Review Required"
            admin_plain = (
                f"A new user has registered and has been granted basic fund-transfer access.\n\n"
                f"Name: {first_name} {last_name}\n"
                f"Username: {username}\n"
                f"Email: {email}\n"
                f"Account Number: {account_number}\n\n"
                f"Review the account or grant High-Value transfer access from the Admin Dashboard: /admin/users/{new_user_id}"
            )
            admin_html = f"""
            <div style="font-family: Arial, sans-serif; max-width: 500px; margin: 0 auto; padding: 24px; background:#ffffff; border:1px solid #eaeaea; border-radius:8px;">
                <h2 style="color:#111827; font-size:18px;">New User Registration</h2>
                <p style="color:#4b5563; font-size:14px;">A new account has registered and been granted <strong>basic fund-transfer access</strong>:</p>
                <table style="width:100%; font-size:13px; color:#374151; border-collapse:collapse; margin:16px 0;">
                    <tr><td style="padding:4px 0; font-weight:600;">Name:</td><td>{first_name} {last_name}</td></tr>
                    <tr><td style="padding:4px 0; font-weight:600;">Username:</td><td>{username}</td></tr>
                    <tr><td style="padding:4px 0; font-weight:600;">Email:</td><td>{email}</td></tr>
                    <tr><td style="padding:4px 0; font-weight:600;">Account #:</td><td>{account_number}</td></tr>
                </table>
                <p style="color:#6b7280; font-size:12px;">Log in to the Admin Dashboard to review the account or enable High-Value transfers.</p>
            </div>
            """
            send_real_email(ADMIN_EMAIL, admin_subject, admin_html, admin_plain)
        except Exception as e:
            # Don't fail registration just because the admin notification email failed
            print(f"[ADMIN NOTIFY ERROR] {e}")

        user_obj = {
            'username': username,
            'name': f"{first_name} {last_name}",
            'email': email,
            'accountType': 'Savings Account',
            'accountNumber': account_number,
            'balance': '25,000.00'
        }
        return jsonify({'success': True, 'user': user_obj, 'message': 'Registration successful. Fund transfers will be enabled once an admin verifies your account.'})
        
    except Exception as e:
        print(f"DB Error: {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE username=%s", (username,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'success': False, 'message': 'Invalid username or password'}), 401

        # Account status is enforced server-side before authentication.
        if (user.get('account_status') or 'ACTIVE') != 'ACTIVE':
            return jsonify({'success': False, 'message': f"Account is {user.get('account_status')}. Please contact the bank administrator."}), 403

        # Check lockout
        now = datetime.now()
        if user['lockout_until']:
            if user['lockout_until'] > now:
                remaining = int((user['lockout_until'] - now).total_seconds())
                print(f"[LOCKOUT BLOCKED] Login attempt rejected for locked user: {username}")
                return jsonify({'success': False, 'error': 'Account locked', 'lockout': True, 'remaining_seconds': remaining, 'message': f'Account locked. Please wait {remaining} seconds.'}), 423
            else:
                cursor.execute("UPDATE users SET failed_attempts=0, lockout_until=NULL WHERE id=%s", (user['id'],))
                conn.commit()
                user['failed_attempts'] = 0
                user['lockout_until'] = None
                print(f"[LOCKOUT EXPIRED] Resetting failed attempts and lockout timestamp for user: {username}")

        # Verify password
        if bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
            # Reset attempts on success
            cursor.execute("UPDATE users SET failed_attempts=0, lockout_until=NULL WHERE id=%s", (user['id'],))
            conn.commit()

            # Do NOT authenticate the server session until the login OTP is verified.
            # Store only a short-lived pre-authentication record in the session.
            flask_session['pending_login_user_id'] = user['id']
            flask_session['pending_login_email'] = user['email']
            flask_session['pending_login_at'] = datetime.now().isoformat()

            user_obj = {
                'username': user['username'],
                'name': f"{user['first_name']} {user['last_name']}",
                'email': user['email'],
                'accountType': 'Savings Account',
                'accountNumber': user['account_number'],
                'balance': str(user['balance'])
            }
            return jsonify({'success': True, 'user': user_obj, 'email': user['email']})
        else:
            # Failed attempt
            attempts = user['failed_attempts'] + 1
            if attempts >= 3:
                lockout_time = now + timedelta(seconds=30)
                cursor.execute("UPDATE users SET failed_attempts=%s, lockout_until=%s WHERE id=%s", (attempts, lockout_time, user['id']))
                conn.commit()
                return jsonify({'success': False, 'error': 'Account locked', 'lockout': True, 'remaining_seconds': 30, 'message': 'Account locked. Please wait 30 seconds.'}), 423
            else:
                cursor.execute("UPDATE users SET failed_attempts=%s WHERE id=%s", (attempts, user['id']))
                conn.commit()
                remaining = 3 - attempts
                msg = f'Incorrect password. {remaining} attempt{"s" if remaining > 1 else ""} remaining.'
                return jsonify({'success': False, 'lockout': False, 'attempts_remaining': remaining, 'message': msg}), 401

    except Exception as e:
        print(f"DB Error: {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@app.route('/api/send-otp', methods=['POST'])
def send_otp():
    data = request.json
    email = data.get('email')
    action = data.get('action')
    frontend_username = data.get('username')
    
    if not email or action not in ('LOGIN','REGISTER','RESET_PASSWORD'):
        return jsonify({'success': False, 'message': 'Invalid OTP request'}), 400
    if action == 'LOGIN':
        pending_email = flask_session.get('pending_login_email')
        if pending_email != email:
            return jsonify({'success': False, 'message': 'Login verification session expired. Please log in again.'}), 403

    otp = str(random.randint(100000, 999999))
    if action == 'LOGIN':
        expires_at = datetime.now() + timedelta(seconds=60)
    else:
        expires_at = datetime.now() + timedelta(seconds=120)
    
    first_name = frontend_username if frontend_username else "Customer"

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT first_name, username FROM users WHERE email=%s", (email,))
        user_record = cursor.fetchone()
        
        if action == 'REGISTER':
            cursor.execute("SELECT id FROM users WHERE username=%s OR email=%s", (frontend_username, email))
            if cursor.fetchone():
                return jsonify({'success': False, 'message': 'Username or email already exists'}), 409
        
        if action == 'RESET_PASSWORD' and not user_record:
            return jsonify({'success': False, 'message': 'No account found with this email address'}), 404

        if user_record:
            if user_record.get('first_name'):
                first_name = user_record['first_name']
            elif user_record.get('username'):
                first_name = user_record['username']

        cursor.execute("INSERT INTO otps (email, otp, action, expires_at) VALUES (%s, %s, %s, %s)", (email, otp, action, expires_at))
        conn.commit()
    except Exception as e:
        print(f"[SEND-OTP ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

    subject, html_body, plain_text = get_email_template(action, first_name, otp)
    send_real_email(email, subject, html_body, plain_text)

    return jsonify({'success': True, 'message': 'OTP sent'})

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    data = request.json
    email = data.get('email')
    otp = data.get('otp')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        requested_action = data.get('action') or 'LOGIN'
        cursor.execute("SELECT * FROM otps WHERE email=%s AND used=FALSE AND action=%s ORDER BY created_at DESC LIMIT 1", (email, requested_action))
        record = cursor.fetchone()

        if not record:
            return jsonify({'success': False, 'message': 'No pending OTP found'})

        if record['expires_at'] < datetime.now():
            return jsonify({'success': False, 'message': 'OTP has expired. Please click Resend OTP.'})

        if record['otp'] == otp:
            cursor.execute("UPDATE otps SET used=TRUE WHERE id=%s", (record['id'],))
            conn.commit()

            action = data.get('action', '') or record.get('action', '')
            if action == 'LOGIN':
                pending_id = flask_session.get('pending_login_user_id')
                pending_email = flask_session.get('pending_login_email')
                if not pending_id or pending_email != email:
                    return jsonify({'success': False, 'message': 'Login verification session expired. Please log in again.'}), 403
                cursor.execute("SELECT * FROM users WHERE id=%s AND email=%s", (pending_id, email))
                login_user = cursor.fetchone()
                if not login_user or (login_user.get('account_status') or 'ACTIVE') != 'ACTIVE':
                    flask_session.pop('pending_login_user_id', None)
                    flask_session.pop('pending_login_email', None)
                    return jsonify({'success': False, 'message': 'Account is not active.'}), 403
                cursor.execute("UPDATE users SET last_login=NOW(), failed_attempts=0, lockout_until=NULL WHERE id=%s", (pending_id,))
                conn.commit()
                flask_session.pop('pending_login_user_id', None)
                flask_session.pop('pending_login_email', None)
                flask_session.pop('pending_login_at', None)
                flask_session['user_id'] = login_user['id']
                flask_session['username'] = login_user['username']
                return jsonify({'success': True, 'user': {
                    'username': login_user['username'],
                    'name': f"{login_user['first_name']} {login_user['last_name']}",
                    'email': login_user['email'], 'accountType': 'Savings Account',
                    'accountNumber': login_user['account_number'],
                    'balance': f"{Decimal(str(login_user['balance'] or 0)):,.2f}",
                    'lastLogin': 'Just now'
                }})
            elif action == 'RESET_PASSWORD':
                flask_session['verified_reset_email'] = email
                flask_session['verified_reset_at'] = datetime.now().isoformat()

            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'Invalid OTP'})
            
    except Exception as e:
        print(f"DB Error: {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        if 'cursor' in locals(): cursor.close()
        if 'conn' in locals(): conn.close()

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    from flask import session as flask_session
    data = request.json
    email = data.get('email')
    new_password = data.get('password')

    if not email or not new_password:
        return jsonify({'success': False, 'message': 'Missing data'}), 400

    # --- Auth check: Flask session (primary) OR DB-verified OTP fallback ---
    session_email = flask_session.get('verified_reset_email')
    is_session_verified = (session_email == email)

    if not is_session_verified:
        # Fallback: check otps table for a verified RESET_PASSWORD OTP within 15 min
        try:
            conn_check = get_db_connection()
            cursor_check = conn_check.cursor(dictionary=True, buffered=True)
            cursor_check.execute(
                "SELECT id FROM otps WHERE email=%s AND action='RESET_PASSWORD' AND used=TRUE AND created_at >= NOW() - INTERVAL 15 MINUTE ORDER BY created_at DESC LIMIT 1",
                (email,)
            )
            otp_fallback = cursor_check.fetchone()
            cursor_check.close()
            conn_check.close()
        except Exception as db_e:
            print(f"[RESET AUTH CHECK ERROR] {db_e}")
            otp_fallback = None

        if not otp_fallback:
            print(f"[RESET REJECTED] No valid session or verified OTP for {email}")
            return jsonify({'success': False, 'message': 'Session expired or unauthorized. Please restart the password reset flow.'}), 403

        print(f"[RESET AUTH] DB-fallback OTP verification passed for {email}")
    else:
        print(f"[RESET AUTH] Session verification passed for {email}")

    try:
        conn = get_db_connection()

        # Fetch the current password hash to check for reuse
        cursor = conn.cursor(dictionary=True, buffered=True)
        cursor.execute("SELECT password_hash FROM users WHERE email=%s", (email,))
        user_row = cursor.fetchone()
        cursor.close()

        if user_row and bcrypt.checkpw(new_password.encode('utf-8'), user_row['password_hash'].encode('utf-8')):
            return jsonify({'success': False, 'error': 'same_password', 'message': 'New password cannot be the same as the old password.'}), 400

        hashed_pw = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())

        cursor = conn.cursor()
        cursor.execute("UPDATE users SET password_hash=%s WHERE email=%s", (hashed_pw, email))
        conn.commit()

        affected = cursor.rowcount
        print(f"[RESET SUCCESS] Password updated for {email}, rows affected: {affected}")
        cursor.close()

        if affected == 0:
            return jsonify({'success': False, 'message': 'No account found with this email address or update failed'}), 404

        # Clear the session flag — single use
        flask_session.pop('verified_reset_email', None)
        flask_session.pop('verified_reset_at', None)

        return jsonify({'success': True, 'message': 'Password reset successful'})
    except Exception as e:
        print(f"[RESET ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        if 'conn' in locals() and conn:
            conn.close()

# --- FUND TRANSFER + TRANSACTION HISTORY ---

@app.route('/api/me', methods=['GET'])
@login_required
def me():
    try:
        user = get_customer_context(flask_session['user_id'])
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        return jsonify({'success': True, 'user': {
            'id': user['id'], 'username': user['username'], 'name': user['name'],
            'email': user['email'], 'accountType': 'Savings Account',
            'accountNumber': user['account_number'], 'balance': user['balance'],
            'midNumber': user['mid_number'], 'lastLogin': user['last_login_display'],
            'fundTransferEnabled': user['fund_transfer_enabled'],
            'accountStatus': user['account_status']
        }})
    except Exception as e:
        print(f"[ME ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500


@app.route('/api/beneficiaries/verify', methods=['POST'])
@login_required
def verify_beneficiary():
    data = request.json or {}
    account_number = (data.get('accountNumber') or '').strip().replace('#', '').replace('-', '')
    confirm = (data.get('confirmAccountNumber') or '').strip().replace('#', '').replace('-', '')
    if not account_number or account_number != confirm:
        return jsonify({'success': False, 'message': 'Account numbers do not match.'}), 400
    import re
    if not re.fullmatch(r'\d{12}', account_number):
        return jsonify({'success': False, 'message': 'Please enter a valid 12-digit account number.'}), 400
    try:
        conn = get_db_connection(); cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, first_name, last_name, account_number, account_status FROM users WHERE account_number=%s", (account_number,))
        user = cursor.fetchone()
        if not user or user['account_status'] != 'ACTIVE':
            return jsonify({'success': False, 'message': 'Active recipient account not found.'}), 404
        if user['id'] == flask_session['user_id']:
            return jsonify({'success': False, 'message': 'You cannot send money to your own account.'}), 400
        return jsonify({'success': True, 'recipient': {
            'firstName': user['first_name'], 'lastName': user['last_name'],
            'accountNumber': user['account_number']
        }})
    except Exception as e:
        print(f"[VERIFY BENEFICIARY ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        try: conn.close()
        except Exception: pass


@app.route('/api/transfer', methods=['POST'])
@login_required
def transfer():
    data = request.json or {}
    to_account = (data.get('toAccount') or '').strip()
    note = (data.get('note') or '').strip()[:255]
    to_account = to_account.replace('#', '').replace('-', '')
    import re
    try:
        amount = Decimal(str(data.get('amount')))
    except (InvalidOperation, TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid amount'}), 400
    if amount < Decimal('1000.00'):
        return jsonify({'success': False, 'message': 'Minimum transfer amount is ₹1,000.'}), 400
    if not to_account:
        return jsonify({'success': False, 'message': 'Recipient account number is required.'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        conn.start_transaction()
        cursor.execute("SELECT * FROM users WHERE id=%s FOR UPDATE", (flask_session['user_id'],))
        sender = cursor.fetchone()
        if not sender or sender['account_status'] != 'ACTIVE':
            conn.rollback(); return jsonify({'success': False, 'message': 'Sender account is not active.'}), 403
        cursor.execute("SELECT fund_transfer FROM user_privileges WHERE user_id=%s", (sender['id'],))
        privilege = cursor.fetchone()
        if not privilege or not privilege['fund_transfer']:
            conn.rollback(); return jsonify({'success': False, 'message': 'Fund transfer is pending admin verification.'}), 403
        cursor.execute("SELECT * FROM users WHERE account_number=%s FOR UPDATE", (to_account,))
        receiver = cursor.fetchone()
        if not receiver or receiver['account_status'] != 'ACTIVE':
            conn.rollback(); return jsonify({'success': False, 'message': 'Active recipient account not found.'}), 404
        if receiver['id'] == sender['id']:
            conn.rollback(); return jsonify({'success': False, 'message': 'You cannot transfer to your own account.'}), 400
        sender_balance = Decimal(str(sender['balance'] or 0))
        if amount > sender_balance:
            conn.rollback(); return jsonify({'success': False, 'message': 'Insufficient available balance.'}), 400
        new_sender = sender_balance - amount
        new_receiver = Decimal(str(receiver['balance'] or 0)) + amount
        reference_id = f"SPX-{uuid.uuid4().hex[:12].upper()}"
        cursor.execute("UPDATE users SET balance=%s WHERE id=%s", (new_sender, sender['id']))
        cursor.execute("UPDATE users SET balance=%s WHERE id=%s", (new_receiver, receiver['id']))
        cursor.execute("""INSERT INTO transactions
            (user_id, type, amount, counterparty_account, counterparty_name, note, description, reference_id, status, balance_after)
            VALUES (%s,'TRANSFER_OUT',%s,%s,%s,%s,%s,%s,'COMPLETED',%s)""",
            (sender['id'], amount, receiver['account_number'], f"{receiver['first_name']} {receiver['last_name']}", note,
             f"Transfer to {receiver['first_name']} {receiver['last_name']}", reference_id, new_sender))
        cursor.execute("""INSERT INTO transactions
            (user_id, type, amount, counterparty_account, counterparty_name, note, description, reference_id, status, balance_after)
            VALUES (%s,'TRANSFER_IN',%s,%s,%s,%s,%s,%s,'COMPLETED',%s)""",
            (receiver['id'], amount, sender['account_number'], f"{sender['first_name']} {sender['last_name']}", note,
             f"Transfer received from {sender['first_name']} {sender['last_name']}", reference_id, new_receiver))
        conn.commit()
        return jsonify({'success': True, 'message': f"₹{amount:,.2f} sent successfully.", 'referenceId': reference_id,
                        'newBalance': f"{new_sender:,.2f}",
                        'recipient': f"{receiver['first_name']} {receiver['last_name']}",
                        'amount': f"{amount:,.2f}", 'date': datetime.now().strftime('%d %b %Y, %I:%M %p')})
    except Exception as e:
        conn.rollback(); print(f"[TRANSFER ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error during transfer.'}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()


@app.route('/api/transactions', methods=['GET'])
@login_required
def transactions():
    limit = max(1, min(request.args.get('limit', 50, type=int), 200))
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""SELECT id,type,amount,counterparty_account,counterparty_name,note,description,reference_id,status,balance_after,created_at
                        FROM transactions WHERE user_id=%s ORDER BY created_at DESC,id DESC LIMIT %s""", (flask_session['user_id'], limit))
        return jsonify({'success': True, 'transactions': transaction_payload(cursor.fetchall())})
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()


@app.route('/api/transactions/statement', methods=['GET'])
@login_required
def transaction_statement():
    duration = request.args.get('duration')
    fy = request.args.get('financial_year')
    fmt = (request.args.get('format') or 'pdf').lower()
    start_raw, end_raw = request.args.get('start_date'), request.args.get('end_date')
    start_date = end_date = None
    try:
        if start_raw and end_raw:
            start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_raw, '%Y-%m-%d').date()
            if end_date < start_date: raise ValueError
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid custom date range.'}), 400
    start_date, end_date = get_statement_range(duration, fy, start_date, end_date)
    rows = fetch_transactions_for_range(flask_session['user_id'], start_date, end_date)
    user = get_customer_context(flask_session['user_id'])
    payload = transaction_payload(rows)
    filename_base = f"SPXBank_Statement_{user['username']}_{start_date}_{end_date}"

    if fmt == 'csv':
        out = io.StringIO(); writer = csv.writer(out)
        writer.writerow(['Date','Description','Type','Sent/Received','Counterparty','Account Number','Amount (INR)','Balance After (INR)','Reference ID','Status'])
        for r,pay in zip(rows,payload):
            writer.writerow([pay['date'], pay['description'] or pay['note'] or pay['typeLabel'], pay['typeLabel'], pay['direction'], pay['counterpartyName'] or '', pay['counterpartyAccount'] or '', pay['amount'], pay['balanceAfter'], pay['referenceId'], pay['status']])
        return Response(out.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={filename_base}.csv'})

    if fmt in ('xlsx','excel'):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title='Statement'
        ws.append(['SPX BANK - Account Statement']); ws.append(['Account Holder', user['name']]); ws.append(['Account Number', user['account_number']]); ws.append(['Period', f'{start_date} to {end_date}']); ws.append([])
        headers=['Date','Description','Type','Sent/Received','Counterparty','Account Number','Amount (INR)','Balance After (INR)','Reference ID','Status']
        ws.append(headers)
        for cell in ws[6]: cell.font=Font(bold=True); cell.fill=PatternFill('solid', fgColor='5C2D91'); cell.font=Font(bold=True, color='FFFFFF')
        for r,pay in zip(rows,payload): ws.append([pay['date'],pay['description'] or pay['note'] or pay['typeLabel'],pay['typeLabel'],pay['direction'],pay['counterpartyName'] or '',pay['counterpartyAccount'] or '',float(r['amount'] or 0),float(r['balance_after'] or 0),pay['referenceId'],pay['status']])
        for col,width in zip(range(1,11),[24,34,18,16,28,24,16,20,20,14]): ws.column_dimensions[chr(64+col)].width=width
        out=io.BytesIO(); wb.save(out); out.seek(0)
        return Response(out.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename_base}.xlsx'})

    buffer=io.BytesIO(); doc=SimpleDocTemplate(buffer,pagesize=letter,topMargin=36,bottomMargin=36)
    styles=getSampleStyleSheet(); elements=[Paragraph('SPX BANK — Account Statement',styles['Title']), Spacer(1,6),
        Paragraph(f"Account Holder: {user['name']}",styles['Normal']), Paragraph(f"Account Number: {user['account_number']}",styles['Normal']),
        Paragraph(f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}",styles['Normal']), Spacer(1,14)]
    data=[['Date','Description','Type','Counterparty','Amount (INR)','Balance (INR)']]
    for r,pay in zip(rows,payload):
        sign='+' if pay['direction']=='IN' else '-'
        data.append([pay['date'], (pay['description'] or pay['note'] or pay['typeLabel'])[:35], pay['typeLabel'], (pay['counterpartyName'] or '-')[:24], f"{sign}₹{float(r['amount'] or 0):,.2f}", f"₹{float(r['balance_after'] or 0):,.2f}"])
    if len(data)==1: elements.append(Paragraph('No transactions found for the selected period.',styles['Normal']))
    else:
        table=Table(data,repeatRows=1,colWidths=[70,105,65,90,75,80]); table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#5C2D91')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#dddddd')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f7f7f7')])]))
        elements.append(table)
    doc.build(elements); buffer.seek(0)
    return Response(buffer.getvalue(),mimetype='application/pdf',headers={'Content-Disposition':f'attachment; filename={filename_base}.pdf'})


@app.route('/api/transactions/statement.pdf', methods=['GET'])
@login_required
def transaction_statement_pdf():
    # Backwards-compatible endpoint: delegates to the same statement logic with PDF output.
    duration=request.args.get('duration','CURRENT_MONTH'); fy=request.args.get('financial_year')
    start_raw=request.args.get('start_date'); end_raw=request.args.get('end_date')
    start_date=end_date=None
    try:
        if start_raw and end_raw:
            start_date=datetime.strptime(start_raw,'%Y-%m-%d').date(); end_date=datetime.strptime(end_raw,'%Y-%m-%d').date()
    except ValueError: return jsonify({'success':False,'message':'Invalid date range'}),400
    start_date,end_date=get_statement_range(duration,fy,start_date,end_date)
    rows=fetch_transactions_for_range(flask_session['user_id'],start_date,end_date)
    user=get_customer_context(flask_session['user_id'])
    buffer=io.BytesIO(); doc=SimpleDocTemplate(buffer,pagesize=letter,topMargin=36,bottomMargin=36); styles=getSampleStyleSheet()
    elements=[Paragraph('SPX BANK — Account Statement',styles['Title']),Spacer(1,6),Paragraph(f"Account Holder: {user['name']}",styles['Normal']),Paragraph(f"Account Number: {user['account_number']}",styles['Normal']),Paragraph(f"Period: {start_date} to {end_date}",styles['Normal']),Spacer(1,14)]
    data=[['Date','Description','Type','Counterparty','Amount (INR)','Balance (INR)']]
    for r in rows:
        incoming=r['type'] in ('TRANSFER_IN','DEPOSIT','ADMIN_CREDIT'); sign='+' if incoming else '-'
        data.append([r['created_at'].strftime('%d %b %Y\n%I:%M %p'),(r.get('description') or r.get('note') or r['type'])[:35],r['type'],(r.get('counterparty_name') or '-')[:24],f"{sign}₹{float(r['amount']):,.2f}",f"₹{float(r['balance_after']):,.2f}"])
    if len(data)>1:
        table=Table(data,repeatRows=1,colWidths=[70,105,65,90,75,80]); table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#5C2D91')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#ddd'))])); elements.append(table)
    else: elements.append(Paragraph('No transactions found for the selected period.',styles['Normal']))
    doc.build(elements); buffer.seek(0)
    return Response(buffer.getvalue(),mimetype='application/pdf',headers={'Content-Disposition':f'attachment; filename=SPXBank_Statement_{user["username"]}.pdf'})


@app.route('/api/profile', methods=['GET'])
@login_required
def get_profile():
    user=get_customer_context(flask_session['user_id'])
    if not user: return jsonify({'success':False,'message':'User not found'}),404
    return jsonify({'success':True,'profile':user})


@app.route('/api/profile', methods=['PUT'])
@login_required
def update_profile():
    data=request.json or {}
    allowed=['date_of_birth','mobile_number','pan','father_name','alternate_email','communication_address','permanent_address','marital_status','religion','category']
    payload={k:(data.get(k) or '').strip() for k in allowed}
    if payload['date_of_birth']:
        try: datetime.strptime(payload['date_of_birth'],'%Y-%m-%d')
        except ValueError: return jsonify({'success':False,'message':'Invalid date of birth'}),400
    if payload['alternate_email'] and '@' not in payload['alternate_email']:
        return jsonify({'success':False,'message':'Enter a valid alternate email address'}),400
    conn=get_db_connection()
    try:
        cursor=conn.cursor()
        cursor.execute("""INSERT INTO add_info (user_id,date_of_birth,mobile_number,pan,father_name,alternate_email,communication_address,permanent_address,marital_status,religion,category)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       ON DUPLICATE KEY UPDATE date_of_birth=VALUES(date_of_birth),mobile_number=VALUES(mobile_number),pan=VALUES(pan),father_name=VALUES(father_name),alternate_email=VALUES(alternate_email),communication_address=VALUES(communication_address),permanent_address=VALUES(permanent_address),marital_status=VALUES(marital_status),religion=VALUES(religion),category=VALUES(category)""",
                    (flask_session['user_id'], payload['date_of_birth'] or None,payload['mobile_number'] or None,payload['pan'] or None,payload['father_name'] or None,payload['alternate_email'] or None,payload['communication_address'] or None,payload['permanent_address'] or None,payload['marital_status'] or None,payload['religion'] or None,payload['category'] or None))
        conn.commit(); return jsonify({'success':True,'message':'Profile information saved successfully.'})
    except Exception as e:
        conn.rollback(); print(f"[PROFILE UPDATE ERROR] {e}"); return jsonify({'success':False,'message':'Unable to save profile'}),500
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()


# --- ADMIN HTML PAGE VIEW ROUTES ---

# ===========================================================================
# CUSTOMER LOAN MANAGEMENT
# ===========================================================================
LOAN_RATES = {
    'PERSONAL': 12.50,
    'HOME': 8.50,
    'EDUCATION': 9.00,
    'GOLD': 10.50,
    'LOAN AGAINST MUTUAL FUND': 9.50,
    'OVERDRAFT AGAINST DEPOSIT': 11.00,
}

def calculate_loan_emi(principal, annual_rate, tenure_months):
    principal = Decimal(str(principal))
    annual_rate = Decimal(str(annual_rate))
    n = int(tenure_months)
    if n <= 0 or principal <= 0:
        raise ValueError('Invalid loan amount or tenure')
    monthly = annual_rate / Decimal('1200')
    if monthly == 0:
        return (principal / Decimal(n)).quantize(Decimal('0.01'))
    factor = (Decimal('1') + monthly) ** n
    emi = principal * monthly * factor / (factor - Decimal('1'))
    return emi.quantize(Decimal('0.01'))

@app.route('/api/loans', methods=['GET'])
@login_required
def customer_get_loans():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM loans WHERE user_id=%s ORDER BY id DESC", (flask_session['user_id'],))
        loans = cursor.fetchall()
        for loan in loans:
            for key in ('amount','interest_rate','emi','monthly_income','existing_emi','outstanding_principal'):
                if loan.get(key) is not None:
                    loan[key] = float(loan[key])
            for key in ('applied_at','approved_at','disbursed_at'):
                if loan.get(key): loan[key] = loan[key].strftime('%Y-%m-%d %H:%M:%S')
            if loan.get('next_emi_date'): loan['next_emi_date'] = loan['next_emi_date'].strftime('%Y-%m-%d')
        return jsonify({'success': True, 'loans': loans})
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()

@app.route('/api/loans/calculate', methods=['POST'])
@login_required
def customer_calculate_loan():
    data = request.json or {}
    loan_type = str(data.get('loan_type','PERSONAL')).strip().upper()
    try:
        amount_str = str(data.get('amount', 0) or 0).replace(',', '')
        amount = Decimal(amount_str)
        tenure = int(data.get('tenure_months', 0) or 0)
    except Exception:
        return jsonify({'success': False, 'message': 'Enter a valid amount and tenure.'}), 400
    rate = Decimal(str(data.get('interest_rate') or LOAN_RATES.get(loan_type, 12.50)))
    if amount < Decimal('10000') or amount > Decimal('10000000') or tenure < 6 or tenure > 360:
        return jsonify({'success': False, 'message': 'Loan amount must be between ₹10,000 and ₹1 Crore and tenure must be between 6 and 360 months.'}), 400
    emi = calculate_loan_emi(amount, rate, tenure)
    total = (emi * tenure).quantize(Decimal('0.01'))
    interest = (total - amount).quantize(Decimal('0.01'))
    return jsonify({'success': True, 'loan_type': loan_type, 'interest_rate': float(rate), 'emi': float(emi), 'total_payable': float(total), 'total_interest': float(interest)})

@app.route('/api/loans/apply', methods=['POST'])
@login_required
def customer_apply_loan():
    data = request.json or {}
    user_id = flask_session['user_id']
    loan_type = str(data.get('loan_type','PERSONAL')).strip().upper()
    try:
        amount_str = str(data.get('amount', 0) or 0).replace(',', '')
        amount = Decimal(amount_str)
        tenure = int(data.get('tenure_months', 0) or 0)
        monthly_income_str = str(data.get('monthly_income', 0) or 0).replace(',', '')
        monthly_income = Decimal(monthly_income_str)
        existing_emi_str = str(data.get('existing_emi', 0) or 0).replace(',', '')
        existing_emi = Decimal(existing_emi_str)
    except Exception:
        return jsonify({'success': False, 'message': 'Enter valid numeric loan details.'}), 400
    purpose = str(data.get('purpose','')).strip()
    employment_type = str(data.get('employment_type','')).strip()
    if loan_type not in LOAN_RATES:
        return jsonify({'success': False, 'message': 'Invalid loan type.'}), 400
    if amount < Decimal('10000') or amount > Decimal('10000000'):
        return jsonify({'success': False, 'message': 'Loan amount must be between ₹10,000 and ₹1 Crore.'}), 400
    if tenure < 6 or tenure > 360:
        return jsonify({'success': False, 'message': 'Tenure must be between 6 and 360 months.'}), 400
    if monthly_income <= 0:
        return jsonify({'success': False, 'message': 'Please enter your monthly income.'}), 400
    if not employment_type:
        return jsonify({'success': False, 'message': 'Please select employment type.'}), 400
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT account_status FROM users WHERE id=%s", (user_id,)); user = cursor.fetchone()
        cursor.execute("SELECT loan_application FROM user_privileges WHERE user_id=%s", (user_id,)); priv = cursor.fetchone() or {}
        if not user or user['account_status'] != 'ACTIVE':
            return jsonify({'success': False, 'message': 'Your account is not active.'}), 403
        if not bool(priv.get('loan_application', True)):
            return jsonify({'success': False, 'message': 'Loan applications are disabled for your account.'}), 403
        cursor.execute("SELECT COUNT(*) AS c FROM loans WHERE user_id=%s AND status IN ('PENDING','APPROVED','ACTIVE')", (user_id,))
        if cursor.fetchone()['c'] >= 3:
            return jsonify({'success': False, 'message': 'You already have three active/pending loan applications.'}), 409
        rate = Decimal(str(LOAN_RATES[loan_type]))
        emi = calculate_loan_emi(amount, rate, tenure)
        
        date_token = datetime.now().strftime("%y%m%d")
        unique_token = random.randint(10000, 99999)
        reference_id = f"SPXPL{date_token}{unique_token}"
        
        cursor.execute("""INSERT INTO loans (user_id, loan_type, amount, interest_rate, tenure_months, emi, status, purpose, employment_type, monthly_income, existing_emi, outstanding_principal, reference_id)
                       VALUES (%s,%s,%s,%s,%s,%s,'PENDING',%s,%s,%s,%s,%s,%s)""",
                       (user_id, loan_type, amount, rate, tenure, emi, purpose or None, employment_type, monthly_income, existing_emi, amount, reference_id))
        loan_id = cursor.lastrowid
        conn.commit()
        return jsonify({'success': True, 'message': 'Loan application submitted successfully.', 'loan_id': loan_id, 'reference_id': reference_id, 'emi': float(emi), 'interest_rate': float(rate), 'status': 'PENDING'})
    except Exception as e:
        conn.rollback(); print(f'[CUSTOMER LOAN APPLY ERROR] {e}')
        return jsonify({'success': False, 'message': 'Unable to submit loan application.'}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()

@app.route('/api/loans/<int:loan_id>/payments', methods=['GET'])
@login_required
def customer_loan_payments(loan_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM loans WHERE id=%s AND user_id=%s", (loan_id, flask_session['user_id']))
        loan = cursor.fetchone()
        if not loan: return jsonify({'success': False, 'message': 'Loan not found.'}), 404
        cursor.execute("SELECT * FROM loan_payments WHERE loan_id=%s ORDER BY paid_at DESC, id DESC", (loan_id,))
        payments = cursor.fetchall()
        for payment in payments:
            for key in ('amount','principal_component','interest_component','balance_after'):
                payment[key] = float(payment[key] or 0)
            if payment.get('paid_at'): payment['paid_at'] = payment['paid_at'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({'success': True, 'payments': payments})
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()

@app.route('/api/loans/<int:loan_id>/pay', methods=['POST'])
@login_required
def customer_pay_loan(loan_id):
    data = request.json or {}
    try: amount = Decimal(str(data.get('amount', 0) or 0))
    except Exception: amount = Decimal('0')
    if amount <= 0: return jsonify({'success': False, 'message': 'Enter a valid payment amount.'}), 400
    user_id = flask_session['user_id']
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True); conn.start_transaction()
        cursor.execute("SELECT * FROM loans WHERE id=%s AND user_id=%s FOR UPDATE", (loan_id, user_id)); loan = cursor.fetchone()
        cursor.execute("SELECT * FROM users WHERE id=%s FOR UPDATE", (user_id,)); user = cursor.fetchone()
        if not loan or loan['status'] != 'ACTIVE': return jsonify({'success': False, 'message': 'Only active loans can be repaid.'}), 400
        outstanding = Decimal(str(loan['outstanding_principal'] or 0)); balance = Decimal(str(user['balance'] or 0))
        if outstanding <= 0: return jsonify({'success': False, 'message': 'This loan is already fully repaid.'}), 400
        if amount > balance: return jsonify({'success': False, 'message': 'Insufficient account balance.'}), 400
        if amount > outstanding: amount = outstanding
        monthly_interest = (outstanding * Decimal(str(loan['interest_rate'] or 0)) / Decimal('1200')).quantize(Decimal('0.01'))
        if amount <= monthly_interest and outstanding > amount:
            return jsonify({'success': False, 'message': f"Payment must be greater than this period's interest of ₹{monthly_interest:,.2f}."}), 400
        interest = min(monthly_interest, amount).quantize(Decimal('0.01'))
        principal = amount - interest
        if principal > outstanding:
            principal = outstanding
            interest = amount - principal
        new_outstanding = (outstanding - principal).quantize(Decimal('0.01')); new_balance = (balance - amount).quantize(Decimal('0.01'))
        ref = 'LNP-' + uuid.uuid4().hex[:12].upper()
        cursor.execute("UPDATE users SET balance=%s WHERE id=%s", (new_balance, user_id))
        next_date = (datetime.now().date().replace(day=1) + timedelta(days=32)).replace(day=5)
        status = 'CLOSED' if new_outstanding <= 0 else 'ACTIVE'
        cursor.execute("UPDATE loans SET outstanding_principal=%s,status=%s,next_emi_date=%s WHERE id=%s", (new_outstanding, status, None if status=='CLOSED' else next_date, loan_id))
        cursor.execute("""INSERT INTO loan_payments (loan_id,user_id,amount,principal_component,interest_component,balance_after,reference_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""", (loan_id,user_id,amount,principal,interest,new_outstanding,ref))
        cursor.execute("""INSERT INTO transactions (user_id,type,amount,counterparty_account,counterparty_name,note,description,reference_id,status,balance_after)
                       VALUES (%s,'LOAN_PAYMENT',%s,NULL,'SPX BANK LOAN',%s,%s,%s,'COMPLETED',%s)""", (user_id,amount,f'Loan #{loan_id}',f'Loan repayment for {loan["loan_type"]}',ref,new_balance))
        conn.commit()
        return jsonify({'success': True, 'message': 'Loan payment successful.', 'referenceId': ref, 'newBalance': f'{new_balance:,.2f}', 'outstanding': f'{new_outstanding:,.2f}', 'status': status})
    except Exception as e:
        conn.rollback(); print(f'[LOAN PAYMENT ERROR] {e}')
        return jsonify({'success': False, 'message': 'Unable to process loan payment.'}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()

@app.route('/admin')
@app.route('/admin/')
def admin_root():
    return redirect('/admin/dashboard')

@app.route('/admin/login')
def admin_login_page():
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
def admin_dashboard_page():
    return render_template('admin_dashboard.html')

@app.route('/admin/users')
def admin_users_page():
    return render_template('admin_users.html')

@app.route('/admin/users/<int:user_id>')
def admin_user_detail_page(user_id):
    return render_template('admin_user_detail.html', user_id=user_id)

@app.route('/admin/loans')
def admin_loans_page():
    return render_template('admin_loans.html')

@app.route('/admin/cards')
def admin_cards_page():
    return render_template('admin_cards.html')

@app.route('/admin/transactions')
def admin_transactions_page():
    return render_template('admin_transactions.html')

@app.route('/admin/audit-logs')
def admin_audit_logs_page():
    return render_template('admin_audit_logs.html')

@app.route('/admin/logout')
def admin_logout_view():
    return redirect('/admin/login')

# --- ADMIN REST API ENDPOINTS ---
@app.route('/api/admin/login', methods=['POST'])
def admin_login_api():
    data = request.json or {}
    email = data.get('email', '').strip()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    login_id = email or username
    
    if (login_id.lower() == ADMIN_EMAIL.lower() or login_id.lower() == 'admin') and password == ADMIN_PASSWORD:
        token = generate_admin_token()
        log_admin_action('ADMIN_LOGIN', target_type='SYSTEM', details='Master Admin logged in successfully')
        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'name': ADMIN_NAME,
                'email': ADMIN_EMAIL,
                'role': 'MASTER_ADMIN'
            },
            'redirect': '/admin/dashboard'
        })
    else:
        log_admin_action('FAILED_ADMIN_LOGIN', target_type='SYSTEM', details=f'Failed admin login attempt for {login_id}', status='FAILED')
        return jsonify({'success': False, 'message': 'Invalid administrator credentials'}), 401

@app.route('/api/admin/dashboard', methods=['GET'])
@require_admin_token
def admin_dashboard_data():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # User Stats
        cursor.execute("SELECT COUNT(*) AS total, SUM(CASE WHEN account_status='ACTIVE' THEN 1 ELSE 0 END) AS active, SUM(CASE WHEN account_status='INACTIVE' THEN 1 ELSE 0 END) AS inactive, SUM(CASE WHEN account_status='LOCKED' OR account_status='SUSPENDED' THEN 1 ELSE 0 END) AS locked FROM users")
        user_stats = cursor.fetchone() or {'total': 0, 'active': 0, 'inactive': 0, 'locked': 0}
        
        # Funds Stats
        cursor.execute("SELECT SUM(balance) AS total_deposits, COUNT(*) AS total_accounts FROM users")
        fund_stats = cursor.fetchone() or {'total_deposits': 0, 'total_accounts': 0}
        
        # Transaction Stats
        cursor.execute("SELECT COUNT(*) AS total_tx, SUM(amount) AS total_volume FROM transactions")
        tx_stats = cursor.fetchone() or {'total_tx': 0, 'total_volume': 0}
        
        # Loan Stats
        cursor.execute("SELECT COUNT(*) AS total, SUM(CASE WHEN status='PENDING' THEN 1 ELSE 0 END) AS pending, SUM(CASE WHEN status='APPROVED' OR status='ACTIVE' THEN 1 ELSE 0 END) AS approved, SUM(CASE WHEN status='REJECTED' THEN 1 ELSE 0 END) AS rejected, SUM(amount) AS total_loan_amt FROM loans")
        loan_stats = cursor.fetchone() or {'total': 0, 'pending': 0, 'approved': 0, 'rejected': 0, 'total_loan_amt': 0}
        
        # Card Stats
        cursor.execute("SELECT COUNT(*) AS total, SUM(CASE WHEN status='ACTIVE' THEN 1 ELSE 0 END) AS active, SUM(CASE WHEN status='BLOCKED' THEN 1 ELSE 0 END) AS blocked, SUM(CASE WHEN status='REQUESTED' OR status='PENDING' THEN 1 ELSE 0 END) AS pending FROM cards")
        card_stats = cursor.fetchone() or {'total': 0, 'active': 0, 'blocked': 0, 'pending': 0}
        
        # Recent Audit Logs (limit 6)
        cursor.execute("SELECT * FROM admin_audit_logs ORDER BY created_at DESC LIMIT 6")
        recent_logs = cursor.fetchall()
        for log in recent_logs:
            if log.get('created_at'):
                log['created_at'] = log['created_at'].strftime('%d %b %Y, %H:%M:%S')
                
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'stats': {
                'users': user_stats,
                'funds': {
                    'total_deposits': float(fund_stats['total_deposits'] or 0),
                    'total_accounts': fund_stats['total_accounts'] or 0
                },
                'transactions': {
                    'total_count': tx_stats['total_tx'] or 0,
                    'total_volume': float(tx_stats['total_volume'] or 0)
                },
                'loans': {
                    'total': loan_stats['total'] or 0,
                    'pending': loan_stats['pending'] or 0,
                    'approved': loan_stats['approved'] or 0,
                    'rejected': loan_stats['rejected'] or 0,
                    'total_amount': float(loan_stats['total_loan_amt'] or 0)
                },
                'cards': card_stats
            },
            'recent_logs': recent_logs
        })
    except Exception as e:
        print(f"[ADMIN DASHBOARD ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch dashboard data'}), 500

@app.route('/api/admin/users', methods=['GET'])
@require_admin_token
def admin_get_users():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        query = "SELECT id, username, email, first_name, last_name, account_number, mid_number, balance, account_status, failed_attempts, created_at, last_login FROM users WHERE 1=1"
        params = []
        
        if search:
            query += " AND (username LIKE %s OR email LIKE %s OR first_name LIKE %s OR last_name LIKE %s OR account_number LIKE %s OR mid_number LIKE %s)"
            pattern = f"%{search}%"
            params.extend([pattern, pattern, pattern, pattern, pattern, pattern])
            
        if status:
            query += " AND account_status = %s"
            params.append(status)
            
        query += " ORDER BY id DESC"
        
        cursor.execute(query, tuple(params))
        users = cursor.fetchall()
        
        for u in users:
            u['balance'] = float(u['balance'] or 0)
            if u.get('created_at'): u['created_at'] = u['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if u.get('last_login'): u['last_login'] = u['last_login'].strftime('%Y-%m-%d %H:%M:%S')
            
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        print(f"[ADMIN USERS GET ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch users'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@require_admin_token
def admin_get_user_detail(user_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404
            
        user['balance'] = float(user['balance'] or 0)
        user.pop('password_hash', None)
        if user.get('created_at'): user['created_at'] = user['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        if user.get('last_login'): user['last_login'] = user['last_login'].strftime('%Y-%m-%d %H:%M:%S')
        
        # User privileges
        cursor.execute("SELECT * FROM user_privileges WHERE user_id=%s", (user_id,))
        privileges = cursor.fetchone()
        if not privileges:
            # Create default privileges
            cursor.execute("INSERT INTO user_privileges (user_id) VALUES (%s)", (user_id,))
            conn.commit()
            cursor.execute("SELECT * FROM user_privileges WHERE user_id=%s", (user_id,))
            privileges = cursor.fetchone()
            
        # User Loans
        cursor.execute("SELECT * FROM loans WHERE user_id=%s ORDER BY id DESC", (user_id,))
        loans = cursor.fetchall()
        for l in loans:
            l['amount'] = float(l['amount'] or 0)
            if l.get('applied_at'): l['applied_at'] = l['applied_at'].strftime('%Y-%m-%d %H:%M:%S')
            
        # User Cards
        cursor.execute("SELECT * FROM cards WHERE user_id=%s ORDER BY id DESC", (user_id,))
        cards = cursor.fetchall()
        for c in cards:
            if c.get('created_at'): c['created_at'] = c['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            
        # Additional customer profile/KYC information
        cursor.execute("SELECT * FROM add_info WHERE user_id=%s", (user_id,))
        add_info = cursor.fetchone() or {}
        if add_info.get('date_of_birth'):
            add_info['date_of_birth'] = add_info['date_of_birth'].strftime('%Y-%m-%d')

        # User Transactions
        cursor.execute("SELECT * FROM transactions WHERE user_id=%s ORDER BY id DESC LIMIT 20", (user_id,))
        transactions = cursor.fetchall()
        for t in transactions:
            t['amount'] = float(t['amount'] or 0)
            if t.get('created_at'): t['created_at'] = t['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'user': user,
            'privileges': privileges,
            'loans': loans,
            'cards': cards,
            'transactions': transactions,
            'add_info': add_info
        })
    except Exception as e:
        print(f"[ADMIN USER DETAIL ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch user details'}), 500

@app.route('/api/admin/users/<int:user_id>/status', methods=['PUT'])
@require_admin_token
def admin_update_user_status(user_id):
    data = request.json or {}
    new_status = data.get('status', '').upper()
    reason = data.get('reason', 'Status updated by Master Admin')
    
    valid_statuses = ['ACTIVE', 'INACTIVE', 'SUSPENDED', 'LOCKED']
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'message': 'Invalid status'}), 400
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT username, email, account_status FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404
            
        prev_status = user['account_status']
        
        if new_status == 'ACTIVE':
            cursor.execute("UPDATE users SET account_status=%s, failed_attempts=0, lockout_until=NULL WHERE id=%s", (new_status, user_id))
        else:
            cursor.execute("UPDATE users SET account_status=%s WHERE id=%s", (new_status, user_id))
            
        conn.commit()
        cursor.close()
        conn.close()
        
        log_admin_action(f'USER_STATUS_{new_status}', target_type='USER', target_id=user_id, details=f"Changed status of user '{user['username']}' to {new_status}. Reason: {reason}", prev_value=prev_status, new_value=new_status)
        
        return jsonify({'success': True, 'message': f"User account status updated to {new_status}"})
    except Exception as e:
        print(f"[ADMIN UPDATE STATUS ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500

@app.route('/api/admin/users/<int:user_id>/privileges', methods=['PUT'])
@require_admin_token
def admin_update_user_privileges(user_id):
    data = request.json or {}
    online_banking = bool(data.get('online_banking', True))
    fund_transfer = bool(data.get('fund_transfer', True))
    card_access = bool(data.get('card_access', True))
    loan_application = bool(data.get('loan_application', True))
    high_value_transfer = bool(data.get('high_value_transfer', False))
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM user_privileges WHERE user_id=%s", (user_id,))
        existing = cursor.fetchone()
        
        if existing:
            cursor.execute("""
            UPDATE user_privileges 
            SET online_banking=%s, fund_transfer=%s, card_access=%s, loan_application=%s, high_value_transfer=%s
            WHERE user_id=%s
            """, (online_banking, fund_transfer, card_access, loan_application, high_value_transfer, user_id))
        else:
            cursor.execute("""
            INSERT INTO user_privileges (user_id, online_banking, fund_transfer, card_access, loan_application, high_value_transfer)
            VALUES (%s, %s, %s, %s, %s, %s)
            """, (user_id, online_banking, fund_transfer, card_access, loan_application, high_value_transfer))
            
        conn.commit()
        cursor.close()
        conn.close()
        
        log_admin_action('UPDATE_USER_PRIVILEGES', target_type='USER', target_id=user_id, details=f"Updated banking privileges for user #{user_id}")
        return jsonify({'success': True, 'message': 'User privileges updated successfully'})
    except Exception as e:
        print(f"[ADMIN UPDATE PRIVILEGES ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500

@app.route('/api/admin/users/<int:user_id>/funds', methods=['POST'])
@require_admin_token
def admin_adjust_user_funds(user_id):
    data = request.json or {}
    amount = float(data.get('amount', 0))
    action_type = data.get('type', 'CREDIT').upper()
    reason = data.get('reason', 'Administrative Balance Adjustment').strip()
    
    if amount <= 0:
        return jsonify({'success': False, 'message': 'Invalid adjustment amount'}), 400
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT username, balance FROM users WHERE id=%s", (user_id,))
        user = cursor.fetchone()
        if not user:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'User not found'}), 404
            
        prev_balance = float(user['balance'] or 0)
        
        if action_type == 'DEBIT' and prev_balance < amount:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': f'Insufficient funds for debit adjustment. Current balance: ₹{prev_balance:,.2f}'}), 400
            
        new_balance = prev_balance + amount if action_type == 'CREDIT' else prev_balance - amount
        
        cursor.execute("UPDATE users SET balance=%s WHERE id=%s", (new_balance, user_id))
        
        ref_id = f"ADM-{uuid.uuid4().hex[:8].upper()}"
        cursor.execute("""
        INSERT INTO transactions (user_id, type, amount, balance_after, description, reference_id, status)
        VALUES (%s, %s, %s, %s, %s, %s, 'COMPLETED')
        """, (user_id, f'ADMIN_{action_type}', amount, new_balance, f"Admin {action_type}: {reason}", ref_id))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        log_admin_action(f'FUND_ADJUSTMENT_{action_type}', target_type='USER', target_id=user_id, details=f"{action_type} ₹{amount:,.2f} to user {user['username']}. Reason: {reason}", prev_value=f"₹{prev_balance:,.2f}", new_value=f"₹{new_balance:,.2f}")
        
        return jsonify({'success': True, 'message': f"Funds {action_type}ED successfully. New Balance: ₹{new_balance:,.2f}", 'new_balance': new_balance})
    except Exception as e:
        print(f"[ADMIN ADJUST FUNDS ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500

@app.route('/api/admin/loans', methods=['GET'])
@require_admin_token
def admin_get_loans():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
        SELECT l.*, u.username, u.first_name, u.last_name, u.email, u.account_number 
        FROM loans l 
        JOIN users u ON l.user_id = u.id 
        ORDER BY l.id DESC
        """)
        loans = cursor.fetchall()
        
        for l in loans:
            for key in ('amount','interest_rate','emi','monthly_income','existing_emi','outstanding_principal'):
                if l.get(key) is not None: l[key] = float(l[key])
            for key in ('applied_at','approved_at','disbursed_at'):
                if l.get(key): l[key] = l[key].strftime('%Y-%m-%d %H:%M:%S')
            if l.get('next_emi_date'): l['next_emi_date'] = l['next_emi_date'].strftime('%Y-%m-%d')
            
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'loans': loans})
    except Exception as e:
        print(f"[ADMIN LOANS GET ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch loans'}), 500

@app.route('/api/admin/loans/<int:loan_id>/action', methods=['PUT'])
@require_admin_token
def admin_loan_action(loan_id):
    data = request.json or {}
    action = data.get('action', '').upper()
    notes = data.get('notes', '')
    valid_actions = ['APPROVE', 'REJECT', 'DISBURSE', 'ACTIVATE', 'CLOSE']
    if action not in valid_actions:
        return jsonify({'success': False, 'message': 'Invalid loan action'}), 400
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True); conn.start_transaction()
        cursor.execute("SELECT * FROM loans WHERE id=%s FOR UPDATE", (loan_id,)); loan = cursor.fetchone()
        if not loan: return jsonify({'success': False, 'message': 'Loan application not found'}), 404
        prev_status = loan['status']; now = datetime.now()
        if action == 'APPROVE':
            if prev_status != 'PENDING': return jsonify({'success': False, 'message': 'Only pending loans can be approved.'}), 400
            new_status = 'APPROVED'
            cursor.execute("UPDATE loans SET status=%s, admin_notes=%s, approved_at=%s WHERE id=%s", (new_status, notes, now, loan_id))
        elif action == 'REJECT':
            if prev_status != 'PENDING': return jsonify({'success': False, 'message': 'Only pending loans can be rejected.'}), 400
            new_status = 'REJECTED'
            cursor.execute("UPDATE loans SET status=%s, admin_notes=%s WHERE id=%s", (new_status, notes, loan_id))
        elif action == 'DISBURSE':
            if prev_status != 'APPROVED': return jsonify({'success': False, 'message': 'Only approved loans can be disbursed.'}), 400
            cursor.execute("SELECT * FROM users WHERE id=%s FOR UPDATE", (loan['user_id'],)); user = cursor.fetchone()
            if not user: return jsonify({'success': False, 'message': 'Customer account not found.'}), 404
            new_balance = (Decimal(str(user['balance'] or 0)) + Decimal(str(loan['amount']))).quantize(Decimal('0.01'))
            ref = 'LND-' + uuid.uuid4().hex[:12].upper()
            next_date = (now.date().replace(day=1) + timedelta(days=32)).replace(day=5)
            cursor.execute("UPDATE users SET balance=%s WHERE id=%s", (new_balance, loan['user_id']))
            cursor.execute("UPDATE loans SET status='ACTIVE', admin_notes=%s, disbursed_at=%s, outstanding_principal=%s, next_emi_date=%s WHERE id=%s", (notes, now, loan['amount'], next_date, loan_id))
            cursor.execute("""INSERT INTO transactions (user_id,type,amount,counterparty_account,counterparty_name,note,description,reference_id,status,balance_after)
                           VALUES (%s,'CREDIT',%s,%s,'SPX BANK','Loan disbursement',%s,%s,'COMPLETED',%s)""", (loan['user_id'], loan['amount'], user['account_number'], f'LOAN_DISBURSEMENT - {loan["loan_type"]} Loan #{loan_id}', ref, new_balance))
            new_status = 'ACTIVE'
        elif action == 'ACTIVATE':
            if prev_status != 'APPROVED': return jsonify({'success': False, 'message': 'Only approved loans can be activated. Use DISBURSE to credit the customer balance.'}), 400
            new_status = 'ACTIVE'
            cursor.execute("UPDATE loans SET status=%s, admin_notes=%s WHERE id=%s", (new_status, notes, loan_id))
        else:
            if prev_status not in ('ACTIVE','APPROVED'): return jsonify({'success': False, 'message': 'Only active/approved loans can be closed.'}), 400
            new_status = 'CLOSED'
            cursor.execute("UPDATE loans SET status=%s, admin_notes=%s WHERE id=%s", (new_status, notes, loan_id))
        conn.commit()
        log_admin_action(f'LOAN_{action}', target_type='LOAN', target_id=loan_id, details=f"Loan #{loan_id} ({loan['loan_type']} ₹{loan['amount']:,.2f}) changed status to {new_status}. Notes: {notes}", prev_value=prev_status, new_value=new_status)
        return jsonify({'success': True, 'message': f'Loan #{loan_id} successfully updated to {new_status}'})
    except Exception as e:
        conn.rollback(); print(f'[ADMIN LOAN ACTION ERROR] {e}')
        return jsonify({'success': False, 'message': 'Server error'}), 500
    finally:
        try: cursor.close()
        except Exception: pass
        conn.close()

@app.route('/api/admin/cards', methods=['GET'])
@require_admin_token
def admin_get_cards():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
        SELECT c.*, u.username, u.first_name, u.last_name, u.email, u.account_number 
        FROM cards c 
        JOIN users u ON c.user_id = u.id 
        ORDER BY c.id DESC
        """)
        cards = cursor.fetchall()
        
        for c in cards:
            if c.get('created_at'): c['created_at'] = c['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if c.get('issued_at'): c['issued_at'] = c['issued_at'].strftime('%Y-%m-%d %H:%M:%S')
            
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'cards': cards})
    except Exception as e:
        print(f"[ADMIN CARDS GET ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch cards'}), 500

@app.route('/api/admin/cards/<int:card_id>/action', methods=['PUT'])
@require_admin_token
def admin_card_action(card_id):
    data = request.json or {}
    action = data.get('action', '').upper()
    
    valid_actions = ['APPROVE', 'ISSUE', 'ACTIVATE', 'BLOCK', 'CANCEL']
    if action not in valid_actions:
        return jsonify({'success': False, 'message': 'Invalid card action'}), 400
        
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM cards WHERE id=%s", (card_id,))
        card = cursor.fetchone()
        if not card:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Card not found'}), 404
            
        prev_status = card['status']
        new_status = action if action in ['ACTIVE', 'BLOCKED', 'CANCELLED'] else ('ISSUED' if action in ['APPROVE', 'ISSUE'] else action)
        
        if action in ['APPROVE', 'ISSUE', 'ACTIVATE']:
            now_dt = datetime.now()
            exp_date = (now_dt + timedelta(days=365*5)).date()
            cursor.execute("UPDATE cards SET status=%s, issued_at=%s, expires_at=%s WHERE id=%s", (new_status, now_dt, exp_date, card_id))
        else:
            cursor.execute("UPDATE cards SET status=%s WHERE id=%s", (new_status, card_id))
            
        conn.commit()
        cursor.close()
        conn.close()
        
        log_admin_action(f'CARD_{action}', target_type='CARD', target_id=card_id, details=f"Card #{card_id} status updated to {new_status}", prev_value=prev_status, new_value=new_status)
        
        return jsonify({'success': True, 'message': f"Card #{card_id} status updated to {new_status}"})
    except Exception as e:
        print(f"[ADMIN CARD ACTION ERROR] {e}")
        return jsonify({'success': False, 'message': 'Server error'}), 500

@app.route('/api/admin/transactions', methods=['GET'])
@require_admin_token
def admin_get_transactions():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
        SELECT t.*, u.username, u.first_name, u.last_name, u.email, u.account_number 
        FROM transactions t 
        JOIN users u ON t.user_id = u.id 
        ORDER BY t.id DESC LIMIT 100
        """)
        transactions = cursor.fetchall()
        
        for t in transactions:
            t['amount'] = float(t['amount'] or 0)
            t['balance_after'] = float(t['balance_after'] or 0)
            if t.get('created_at'): t['created_at'] = t['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'transactions': transactions})
    except Exception as e:
        print(f"[ADMIN TX GET ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch transactions'}), 500

@app.route('/api/admin/audit-logs', methods=['GET'])
@require_admin_token
def admin_get_audit_logs():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM admin_audit_logs ORDER BY id DESC LIMIT 200")
        logs = cursor.fetchall()
        
        for l in logs:
            if l.get('created_at'): l['created_at'] = l['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'audit_logs': logs})
    except Exception as e:
        print(f"[ADMIN AUDIT LOGS ERROR] {e}")
        return jsonify({'success': False, 'message': 'Failed to fetch audit logs'}), 500

@app.route('/api/admin/profile', methods=['GET'])
@require_admin_token
def admin_get_profile():
    return jsonify({
        'success': True,
        'profile': {
            'name': ADMIN_NAME,
            'email': ADMIN_EMAIL,
            'role': 'MASTER_ADMIN',
            'system_status': 'ACTIVE',
            'permissions': ['ALL_PERMISSIONS', 'MASTER_CONTROL']
        }
    })



if __name__ == '__main__':
    app.run(port=5000, debug=True)
